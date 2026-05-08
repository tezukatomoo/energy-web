import io
import re
import unicodedata
from typing import Dict, Optional, Tuple
from datetime import datetime

import streamlit as st
import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT


# =========================================================
# ページ設定
# =========================================================
st.set_page_config(
    page_title="東京都環境計画書 集計ツール",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Streamlitのデフォルトヘッダー・フッター・余白を削除
st.markdown("""
<style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    .block-container {
        padding-top: 0rem;
        padding-bottom: 0rem;
        max-width: 100%;
    }
    [data-testid="stAppViewContainer"] > .main {
        padding-top: 0rem;
    }
    [data-testid="stHeader"] {
        display: none;
    }
    .stApp > header {
        display: none;
    }
</style>
""", unsafe_allow_html=True)

# カスタムCSS
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;700&display=swap');
    
    * {
        font-family: 'Noto Sans JP', sans-serif;
    }
    
    .main {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
    .stApp {
        background: linear-gradient(to bottom, #f8f9fa 0%, #e9ecef 100%);
    }
    div[data-testid="stMetricValue"] {
        font-size: 28px;
        font-weight: 700;
    }
    .upload-box {
        background: white;
        padding: 2rem;
        border-radius: 15px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        margin-bottom: 1.5rem;
    }
    .result-box {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 2rem;
        border-radius: 15px;
        margin: 1rem 0;
    }
    h1 {
        color: #667eea;
        font-weight: 700;
        text-align: center;
        padding: 1rem 0;
    }
    h2, h3 {
        color: #764ba2;
        font-weight: 600;
    }
    .stButton>button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        font-size: 16px;
        transition: all 0.3s ease;
    }
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 12px rgba(102, 126, 234, 0.4);
    }
    .stDownloadButton>button {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        font-size: 16px;
        width: 100%;
    }
    .stDownloadButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 12px rgba(245, 87, 108, 0.4);
    }
    div[data-testid="stExpander"] {
        background: white;
        border-radius: 10px;
        border: 2px solid #e9ecef;
    }
    /* 入力欄のスタイル */
    .stTextInput>div>div>input {
        background-color: white !important;
        color: #333 !important;
        border: 2px solid #667eea !important;
        border-radius: 8px !important;
        padding: 12px !important;
        font-size: 16px !important;
    }
    .stTextInput>div>div>input:focus {
        border-color: #764ba2 !important;
        box-shadow: 0 0 0 3px rgba(118, 75, 162, 0.1) !important;
    }
    /* ファイルアップロード欄のスタイル */
    [data-testid="stFileUploader"] {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        border: 3px dashed #667eea;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    [data-testid="stFileUploader"]:hover {
        border-color: #764ba2;
        background: #f8f9ff;
    }
    [data-testid="stFileUploader"] section {
        border: none !important;
    }
    /* ログインボックス */
    .login-box {
        background: white;
        padding: 3rem;
        border-radius: 20px;
        box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        margin: 2rem 0;
    }
</style>
""", unsafe_allow_html=True)


# =========================================================
# 簡易パスワード認証（試用用）
# =========================================================
PASSWORD = "energy2026"

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("""
            <div style='text-align: center; margin-top: 10rem; margin-bottom: 3rem;'>
                <h1 style='font-size: 2.5rem; margin-bottom: 0.5rem; color: #667eea;'>⚡ 東京都環境計画書 ⚡</h1>
                <h2 style='color: #764ba2; font-weight: 400; font-size: 1.3rem;'>消費電力量集計ツール</h2>
            </div>
        """, unsafe_allow_html=True)
        
        pw = st.text_input("🔐 パスワード", type="password", key="password_input", placeholder="パスワードを入力")
        if st.button("ログイン", use_container_width=True, key="login_button"):
            if pw == PASSWORD:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("❌ パスワードが違います")
    st.stop()


# =========================================================
# メイン画面（ログイン後）
# =========================================================


# =========================================================
# タイプキー抽出
# =========================================================
def extract_type_key_from_filename(name: str) -> str:
    s = unicodedata.normalize("NFKC", name).strip()
    s = s.replace("／", "/")
    if "/" in s:
        s = s.split("/")[-1]
    if s.lower().endswith(".pdf"):
        s = s[:-4]
    return s.strip()


def extract_type_key_from_label(label: str) -> str:
    s = unicodedata.normalize("NFKC", str(label)).strip()
    s = s.replace("／", "/")
    if "/" in s:
        s = s.split("/")[-1]
    return s.strip()


# =========================================================
# PDFから消費電力量[kWh]を抽出（専用部）
# =========================================================
def extract_kwh_from_pdf_bytes(pdf_bytes: bytes) -> Optional[int]:
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            page = pdf.pages[-1]
            raw = page.extract_text() or ""
    except Exception:
        return None

    raw = unicodedata.normalize("NFKC", raw).replace("ｋＷｈ", "kWh")
    lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]

    for i, ln in enumerate(lines):
        if "消費電力量" in ln and "kWh" in ln:
            for j in range(1, 4):
                if i + j < len(lines):
                    m = re.search(r"([0-9]{3,}(?:,[0-9]{3})*)", lines[i + j])
                    if m:
                        return int(m.group(1).replace(",", ""))
            m = re.search(r"([0-9]{3,}(?:,[0-9]{3})*)", ln)
            if m:
                return int(m.group(1).replace(",", ""))
    return None


# =========================================================
# 共用部PDFから消費電力量を抽出（3ページ目）
# =========================================================
def extract_common_area_energy(pdf_bytes: bytes) -> Tuple[Optional[float], Optional[float], Optional[float], list]:
    """共用部PDFから「建物全体」「太陽光削減量」「実消費電力」(MWh) を抽出。

    対応フォーマット:
      - 新 (Ver.3.10 2026.04 以降): 4ページ目に「二次エネルギー消費量計算結果」。
        太陽光発電は正の値（例: 5.33）で表示される。
      - 旧: 3ページ目に「二次エネルギー消費量計算結果」。
        太陽光発電はマイナス符号付き（例: -5.78）で表示される。

    solar_reduction は内部的に常に正の「削減量」として保持し、
    actual_consumption = building_total + solar_reduction を返す。
    """
    debug_info = []

    raw = None
    page_used = None
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            debug_info.append(f"PDFページ数: {len(pdf.pages)}ページ")
            # 4ページ目→3ページ目の順に「二次エネルギー消費量計算結果」を探す（新旧両対応）
            for idx in (3, 2):
                if idx < len(pdf.pages):
                    txt = pdf.pages[idx].extract_text() or ""
                    txt_norm = unicodedata.normalize("NFKC", txt)
                    if (
                        "二次エネルギー消費量計算結果" in txt_norm
                        and "建物全体" in txt_norm
                    ):
                        raw = txt_norm
                        page_used = idx + 1
                        debug_info.append(
                            f"✓ {page_used}ページ目から「二次エネルギー消費量計算結果」を検出: {len(raw)}文字"
                        )
                        break
            if raw is None:
                debug_info.append(
                    "❌ 「二次エネルギー消費量計算結果」が3〜4ページ目に見つかりません"
                )
                return None, None, None, debug_info
    except Exception as e:
        debug_info.append(f"❌ PDF読み込みエラー: {str(e)}")
        return None, None, None, debug_info

    lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
    debug_info.append(f"抽出行数: {len(lines)}行")

    section_start_idx = None
    for i, ln in enumerate(lines):
        if "二次エネルギー消費量計算結果" in ln:
            section_start_idx = i
            debug_info.append(f"✓ セクション発見(行{i}): {ln[:50]}")
            break

    if section_start_idx is None:
        debug_info.append("❌ 二次エネルギー消費量計算結果セクションが見つかりません")
        return None, None, None, debug_info

    building_total = None
    solar_reduction = None
    building_idx = None

    for i in range(section_start_idx, min(section_start_idx + 30, len(lines))):
        ln = lines[i]
        # 「建物全体（延床面積あたり）」は除外
        if "建物全体" in ln and "延床" not in ln and building_total is None:
            building_idx = i
            for offset in range(0, 5):
                if i + offset < len(lines):
                    search_line = lines[i + offset]
                    match = re.search(r"(\d+\.\d+)", search_line)
                    if match:
                        building_total = float(match.group(1))
                        debug_info.append(f"✓ 建物全体の値: {building_total} MWh")
                        break
            if building_total is not None:
                break

    if building_idx is not None:
        for i in range(max(section_start_idx, building_idx - 20), building_idx):
            ln = lines[i]
            if "太陽光" in ln or "PV" in ln:
                # マイナス符号あり/なし両対応。値は常に正の「削減量」として保持する
                for offset in range(0, 4):
                    if i + offset < len(lines):
                        search_line = lines[i + offset]
                        match = re.search(r"(-?\d+\.\d+)", search_line)
                        if match:
                            solar_reduction = abs(float(match.group(1)))
                            debug_info.append(
                                f"✓ 太陽光削減量: {solar_reduction} MWh（符号は除去して正値で保持）"
                            )
                            break
                if solar_reduction is not None:
                    break

    if building_total is not None and solar_reduction is not None:
        # 建物全体は太陽光削減後の値。実消費 = 建物全体 + 太陽光削減量
        actual_consumption = building_total + solar_reduction
        debug_info.append(
            f"✓ 計算完了: {building_total} + {solar_reduction} = {actual_consumption} MWh"
        )
        return building_total, solar_reduction, actual_consumption, debug_info

    return building_total, solar_reduction, None, debug_info


# =========================================================
# 住戸リストCSVの列検出
# =========================================================
def detect_unitlist_columns(df: pd.DataFrame):
    col_row = next(c for c in df.columns if "行" in c)
    col_num = next(c for c in df.columns if ("住戸" in c and "番号" in c))
    candidates = [
        c for c in df.columns
        if ("住宅タイプ" in c) or ("タイプ" in c and "名称" in c)
    ]
    if not candidates:
        raise RuntimeError("『住宅タイプの名称』列が見つかりません")
    return col_row, col_num, candidates[0]


# =========================================================
# PDF出力機能
# =========================================================
def build_pdf_report(
    unit_list: pd.DataFrame,
    project_name: str,
    common_area_mwh: Optional[float] = None,
    building_total: Optional[float] = None,
    solar_reduction: Optional[float] = None
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )
    
    # 日本語フォント設定（OS別TTFを順に試し、最後はreportlab内蔵CIDフォントに必ずフォールバック）
    font_name = None
    for ttf_path, idx in [
        ('C:\\Windows\\Fonts\\msgothic.ttc', 0),               # Windows
        ('/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc', 0),  # macOS
        ('/usr/share/fonts/truetype/fonts-japanese-gothic.ttf', None),
        ('/usr/share/fonts/opentype/ipafont-gothic/ipagp.ttf', None),
    ]:
        try:
            if idx is not None:
                pdfmetrics.registerFont(TTFont('Japanese', ttf_path, subfontIndex=idx))
            else:
                pdfmetrics.registerFont(TTFont('Japanese', ttf_path))
            font_name = 'Japanese'
            break
        except Exception:
            continue
    if font_name is None:
        # Render等の最小環境向け: reportlab同梱のCID日本語フォント（追加パッケージ不要）
        try:
            pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
            font_name = 'HeiseiKakuGo-W5'
        except Exception:
            font_name = 'Courier'
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=font_name,
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontName=font_name,
        fontSize=14,
        spaceAfter=10
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=10
    )
    
    elements = []
    
    elements.append(Paragraph(project_name, title_style))
    elements.append(Paragraph(f"作成日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}", normal_style))
    elements.append(Spacer(1, 10*mm))
    
    elements.append(Paragraph("集計結果サマリー", heading_style))
    
    total_private_kwh = int(unit_list["消費電力量[kWh]"].sum())
    summary_data = [["専用部合計消費電力量", f"{total_private_kwh:,} kWh"]]
    
    if common_area_mwh:
        common_kwh = int(common_area_mwh * 1000)
        grand_total = total_private_kwh + common_kwh
        summary_data.extend([
            ["共用部消費電力量", f"{common_kwh:,} kWh"],
            ["建物全体消費電力量", f"{grand_total:,} kWh"]
        ])
    
    summary_table = Table(summary_data, colWidths=[80*mm, 80*mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
        ('BACKGROUND', (0, -1), (-1, -1), colors.yellow),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 10*mm))
    
    if common_area_mwh and building_total is not None and solar_reduction is not None:
        elements.append(Paragraph("共用部消費電力量の計算内訳", heading_style))
        
        common_detail_data = [
            ["項目", "値"],
            ["建物全体（太陽光削減後）", f"{building_total:.2f} MWh"],
            ["太陽光削減量", f"{solar_reduction:.2f} MWh"],
            ["実際の消費電力（太陽光削減前）", f"{common_area_mwh:.2f} MWh"],
            ["", f"= {common_area_mwh * 1000:.0f} kWh"]
        ]
        
        common_detail_table = Table(common_detail_data, colWidths=[80*mm, 80*mm])
        common_detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('BACKGROUND', (0, 3), (-1, 3), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(common_detail_table)
        elements.append(Spacer(1, 5*mm))
        
        calc_text = f"計算式: {building_total:.2f} + {solar_reduction:.2f} = {common_area_mwh:.2f} MWh"
        elements.append(Paragraph(calc_text, normal_style))
        elements.append(Spacer(1, 10*mm))
    
    elements.append(Paragraph("タイプ別集計", heading_style))
    
    type_summary = (
        unit_list
        .groupby("タイプ", as_index=False)
        .agg(戸数=("住戸の番号", "count"), 合計消費電力量=("消費電力量[kWh]", "sum"))
    )
    type_summary["1住戸あたり"] = (type_summary["合計消費電力量"] / type_summary["戸数"]).round(0).astype(int)
    
    type_data = [["タイプ", "戸数", "1住戸あたり[kWh]", "合計[kWh]"]]
    for _, row in type_summary.sort_values("タイプ").iterrows():
        type_data.append([
            str(row["タイプ"]),
            f"{int(row['戸数'])}",
            f"{int(row['1住戸あたり']):,}",
            f"{int(row['合計消費電力量']):,}"
        ])
    
    type_table = Table(type_data, colWidths=[40*mm, 30*mm, 45*mm, 45*mm])
    type_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(type_table)
    elements.append(PageBreak())
    
    elements.append(Paragraph("住戸別詳細", heading_style))
    
    detail_data = [["行番号", "住戸番号", "タイプ", "消費電力量[kWh]"]]
    for _, row in unit_list.iterrows():
        detail_data.append([
            str(row["行番号"]),
            str(row["住戸の番号"]),
            str(row["タイプ"]),
            f"{int(row['消費電力量[kWh]']) if pd.notna(row['消費電力量[kWh]']) else '-':,}" if pd.notna(row['消費電力量[kWh]']) else "-"
        ])
    
    detail_table = Table(detail_data, colWidths=[25*mm, 35*mm, 40*mm, 60*mm])
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (2, -1), 'CENTER'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(detail_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


# =========================================================
# Excel作成
# =========================================================
def build_standard_excel(
    unit_list: pd.DataFrame, 
    project_name: str,
    common_area_mwh: Optional[float] = None
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "集計"

    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="667EEA")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    title_fill = PatternFill("solid", fgColor="764BA2")
    common_fill = PatternFill("solid", fgColor="E8DAEF")
    grand_fill = PatternFill("solid", fgColor="F5576C")
    bold = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    t = ws.cell(row=1, column=1)
    t.value = project_name
    t.font = Font(bold=True, size=16, color="FFFFFF")
    t.alignment = center
    t.fill = title_fill

    left_headers = ["行番号", "住戸の番号", "タイプ", "消費電力量[kWh]"]
    for c, h in enumerate(left_headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for i, row in unit_list.iterrows():
        r = i + 3
        ws.cell(row=r, column=1, value=row["行番号"]).border = border
        ws.cell(row=r, column=2, value=row["住戸の番号"]).border = border
        ws.cell(row=r, column=3, value=row["タイプ"]).border = border
        ws.cell(row=r, column=4, value=row["消費電力量[kWh]"]).border = border

        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=2).alignment = right
        ws.cell(row=r, column=3).alignment = center
        ws.cell(row=r, column=4).alignment = right

    total_units = int(unit_list["住戸の番号"].nunique())
    total_kwh = int(unit_list["消費電力量[kWh]"].sum())
    sum_row = len(unit_list) + 3

    ws.cell(row=sum_row, column=1, value="専用部合計住戸数").fill = total_fill
    ws.cell(row=sum_row, column=2, value=total_units).fill = total_fill
    ws.cell(row=sum_row, column=3, value="専用部合計消費電力量[kWh]").fill = total_fill
    ws.cell(row=sum_row, column=4, value=total_kwh).fill = total_fill

    for c in range(1, 5):
        ws.cell(row=sum_row, column=c).font = Font(bold=True)
        ws.cell(row=sum_row, column=c).border = border

    if common_area_mwh is not None:
        common_kwh = int(common_area_mwh * 1000)
        sum_row += 1
        ws.cell(row=sum_row, column=3, value="共用部消費電力量[kWh]").fill = common_fill
        ws.cell(row=sum_row, column=4, value=common_kwh).fill = common_fill
        ws.cell(row=sum_row, column=3).font = Font(bold=True)
        ws.cell(row=sum_row, column=4).font = Font(bold=True)
        ws.cell(row=sum_row, column=3).border = border
        ws.cell(row=sum_row, column=4).border = border
        ws.cell(row=sum_row, column=4).alignment = right

        grand_total = total_kwh + common_kwh
        sum_row += 1
        ws.cell(row=sum_row, column=3, value="建物全体消費電力量[kWh]").fill = grand_fill
        ws.cell(row=sum_row, column=4, value=grand_total).fill = grand_fill
        ws.cell(row=sum_row, column=3).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=sum_row, column=4).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=sum_row, column=3).border = border
        ws.cell(row=sum_row, column=4).border = border
        ws.cell(row=sum_row, column=4).alignment = right

    ts = (
        unit_list
        .groupby("タイプ", as_index=False)
        .agg(戸数=("住戸の番号", "count"), 合計消費電力量_kWh=("消費電力量[kWh]", "sum"))
    )
    ts["kwh_per_unit"] = (ts["合計消費電力量_kWh"] / ts["戸数"]).round(0).astype(int)

    right_headers = ["タイプ", "戸数", "1住戸あたり消費電力量[kWh]", "合計消費電力量[kWh]"]
    for c, h in enumerate(right_headers, start=6):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    r0 = 3
    for i, row in ts.sort_values("タイプ").iterrows():
        ws.cell(row=r0, column=6, value=row["タイプ"]).border = border
        ws.cell(row=r0, column=7, value=int(row["戸数"])).border = border
        ws.cell(row=r0, column=8, value=int(row["kwh_per_unit"])).border = border
        ws.cell(row=r0, column=9, value=int(row["合計消費電力量_kWh"])).border = border

        for c in range(6, 10):
            ws.cell(row=r0, column=c).alignment = right if c >= 7 else center
        r0 += 1
    
    sum_units = int(ts["戸数"].sum())
    sum_kwh = int(ts["合計消費電力量_kWh"].sum())
    r0 += 1

    ws.cell(row=r0, column=6, value="専用部合計住戸数").fill = total_fill
    ws.cell(row=r0, column=7, value=sum_units).fill = total_fill
    ws.cell(row=r0, column=6).font = Font(bold=True)
    ws.cell(row=r0, column=7).font = Font(bold=True)
    ws.cell(row=r0, column=6).border = border
    ws.cell(row=r0, column=7).border = border
    ws.cell(row=r0, column=6).alignment = center
    ws.cell(row=r0, column=7).alignment = right

    r0 += 1
    ws.cell(row=r0, column=6, value="専用部合計消費電力量[kWh]").fill = total_fill
    ws.cell(row=r0, column=7, value=sum_kwh).fill = total_fill
    ws.cell(row=r0, column=6).font = Font(bold=True)
    ws.cell(row=r0, column=7).font = Font(bold=True)
    ws.cell(row=r0, column=6).border = border
    ws.cell(row=r0, column=7).border = border
    ws.cell(row=r0, column=6).alignment = center
    ws.cell(row=r0, column=7).alignment = right

    if common_area_mwh is not None:
        common_kwh = int(common_area_mwh * 1000)
        
        r0 += 1
        ws.cell(row=r0, column=6, value="共用部消費電力量[kWh]").fill = common_fill
        ws.cell(row=r0, column=7, value=common_kwh).fill = common_fill
        ws.cell(row=r0, column=6).font = Font(bold=True)
        ws.cell(row=r0, column=7).font = Font(bold=True)
        ws.cell(row=r0, column=6).border = border
        ws.cell(row=r0, column=7).border = border
        ws.cell(row=r0, column=6).alignment = center
        ws.cell(row=r0, column=7).alignment = right

        grand_total = sum_kwh + common_kwh
        r0 += 1
        ws.cell(row=r0, column=6, value="建物全体消費電力量[kWh]").fill = grand_fill
        ws.cell(row=r0, column=7, value=grand_total).fill = grand_fill
        ws.cell(row=r0, column=6).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=r0, column=7).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=r0, column=6).border = border
        ws.cell(row=r0, column=7).border = border
        ws.cell(row=r0, column=6).alignment = center
        ws.cell(row=r0, column=7).alignment = right

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 26
    ws.column_dimensions["I"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# =========================================================
# メイン画面(ログイン後)
# =========================================================
st.markdown("<h1 style='margin-top: 2rem;'>⚡ 東京都環境計画書 消費電力量集計ツール</h1>", unsafe_allow_html=True)
st.markdown("<p style='text-align: center; color: #666; font-size: 18px; margin-bottom: 2rem;'>専用部・共用部の消費電力量を簡単に集計</p>", unsafe_allow_html=True)

col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    st.markdown("### 📝 物件情報")
    project_name = st.text_input("物件名", value="（仮称）〇〇計画 新築工事", label_visibility="collapsed")

st.markdown("<div style='margin: 2rem 0;'></div>", unsafe_allow_html=True)

col1, col2, col3 = st.columns(3)

with col1:
    st.markdown("### 📄 住戸リスト")
    csv_file = st.file_uploader("CSVファイル", type=["csv"], label_visibility="collapsed")

with col2:
    st.markdown("### 🏠 専用部PDF")
    pdf_files = st.file_uploader("タイプ別PDF（複数可）", type=["pdf"], accept_multiple_files=True, label_visibility="collapsed")

with col3:
    st.markdown("### 🏢 共用部PDF")
    common_pdf = st.file_uploader("共用部PDF（1ファイル）", type=["pdf"], key="common_pdf", label_visibility="collapsed")

col1, col2, col3 = st.columns([1, 1, 1])
with col2:
    if st.button("🚀 集計実行", use_container_width=True):
        if not csv_file or not pdf_files:
            st.error("❌ CSVと専用部PDFを両方アップロードしてください")
        else:
            with st.spinner("⏳ 処理中..."):
                type_kwh: Dict[str, Optional[int]] = {}
                rows = []

                for f in pdf_files:
                    kwh = extract_kwh_from_pdf_bytes(f.read())
                    tkey = extract_type_key_from_filename(f.name)
                    rows.append({"PDF名": f.name, "タイプ": tkey, "kWh": kwh})
                    type_kwh[tkey] = kwh

                st.markdown("<div class='result-box'>", unsafe_allow_html=True)
                st.markdown("### ✅ 専用部PDF抽出結果")
                st.dataframe(pd.DataFrame(rows), use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)

                common_area_mwh = None
                building_total_value = None
                solar_reduction_value = None
                
                if common_pdf:
                    building_total, solar_reduction, actual_consumption, debug_info = extract_common_area_energy(common_pdf.read())
                    
                    with st.expander("🔍 抽出デバッグ情報", expanded=False):
                        for info in debug_info:
                            st.text(info)
                    
                    if actual_consumption is not None:
                        st.success("✅ 共用部消費電力量を抽出しました")
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("🏢 建物全体（太陽光削減後）", f"{building_total:.2f} MWh")
                        with col2:
                            st.metric("☀️ 太陽光削減量", f"{solar_reduction:.2f} MWh")
                        with col3:
                            st.metric("⚡ 実際の消費電力", f"{actual_consumption:.2f} MWh", delta=f"{actual_consumption * 1000:.0f} kWh")
                        common_area_mwh = actual_consumption
                        building_total_value = building_total
                        solar_reduction_value = solar_reduction
                    else:
                        st.error("⚠️ 共用部PDFから値を抽出できませんでした")

                units = None
                for enc in ("utf-8-sig", "cp932", "utf-8"):
                    try:
                        units = pd.read_csv(csv_file, encoding=enc)
                        break
                    except Exception:
                        continue

                if units is None:
                    st.error("❌ CSVを読み込めませんでした")
                else:
                    col_row, col_num, col_type = detect_unitlist_columns(units)
                    units["タイプ"] = units[col_type].apply(extract_type_key_from_label)
                    units["消費電力量[kWh]"] = units["タイプ"].map(type_kwh)
                    unit_list = units[[col_row, col_num, "タイプ", "消費電力量[kWh]"]]
                    unit_list.columns = ["行番号", "住戸の番号", "タイプ", "消費電力量[kWh]"]

                    with st.expander("📋 住戸別マッピング（先頭50行）", expanded=False):
                        st.dataframe(unit_list.head(50), use_container_width=True)

                    missing = unit_list[unit_list["消費電力量[kWh]"].isna()]
                    if not missing.empty:
                        st.warning("⚠️ kWhが取得できていないタイプがあります")
                        st.dataframe(missing["タイプ"].value_counts())

                    st.markdown("<div class='result-box'>", unsafe_allow_html=True)
                    st.markdown("### 📊 集計結果")
                    
                    col1, col2, col3 = st.columns(3)
                    total_private = int(unit_list["消費電力量[kWh]"].sum())
                    
                    with col1:
                        st.metric("🏠 専用部合計", f"{total_private:,} kWh")
                    
                    if common_area_mwh:
                        common_kwh = int(common_area_mwh * 1000)
                        with col2:
                            st.metric("🏢 共用部", f"{common_kwh:,} kWh")
                        with col3:
                            st.metric("🏗️ 建物全体", f"{total_private + common_kwh:,} kWh")
                    
                    st.markdown("</div>", unsafe_allow_html=True)

                    st.markdown("### 💾 ダウンロード")
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        excel = build_standard_excel(unit_list, project_name, common_area_mwh)
                        st.download_button("📊 Excelダウンロード", data=excel, file_name=f"{project_name}_消費電力量集計.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    
                    with col2:
                        pdf_report = build_pdf_report(unit_list, project_name, common_area_mwh, building_total_value, solar_reduction_value)
                        st.download_button("📄 PDF出力", data=pdf_report, file_name=f"{project_name}_消費電力量集計.pdf", mime="application/pdf", use_container_width=True)
                    
                    with col3:
                        st.info("💡 PDFをダウンロードして印刷できます")